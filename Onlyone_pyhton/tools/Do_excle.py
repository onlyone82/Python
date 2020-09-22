from openpyxl import load_workbook
from tools.read_config import Readconfig
from tools import project_path
from tools.Get_Testdata import Get_data
class Do_excle:
    @classmethod
    def get_data(cls,file_name):
        tel=getattr(Get_data,'tel')
        wb=load_workbook(file_name)
        mode=eval(Readconfig.get_config(project_path.config_path, 'MODE', 'mode'))
        test_data=[]
        for key in mode:
            sheet=wb[key]
            if mode[key]=="all":
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['case_id']=sheet.cell(i,1).value
                    row_data['url']=sheet.cell(i,2).value
                    if sheet.cell(i,3).value.find("${tel_1}")!=-1:
                        row_data["data"]=sheet.cell(i,3).value.replace("${tel_1}",str(getattr(Get_data,'login_tel')))
                    elif sheet.cell(i, 3).value.find("${tel}") != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace("${tel}", str(tel+1))
                    elif sheet.cell(i, 3).value.find("$loan_member_id") != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace("$loan_member_id",str(getattr(Get_data,"loan_member_id")))
                    elif sheet.cell(i, 3).value.find("$normal_tel") != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace("$normal_tel", str(getattr(Get_data,"normal_tel")))
                    elif sheet.cell(i, 3).value.find("$memberID") != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace("$memberID", str(getattr(Get_data,"memberID")))
                    else:
                        row_data['data'] = sheet.cell(i, 3).value
                    row_data['title']=sheet.cell(i,4).value
                    row_data['http_method']=sheet.cell(i,5).value
                    row_data['expected']=sheet.cell(i,6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
                    cls.update_tel(tel+1,file_name,'init')
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    #做手机号的替换
                    if sheet.cell(case_id+1, 3).value.find("${tel_1}") != -1:
                        row_data["data"] = sheet.cell(case_id+1,3).value.replace("${tel}",str(tel))
                    elif sheet.cell(case_id+1, 3).value.find("${tel}") != -1:
                        row_data['data'] = sheet.cell(case_id+1,3).value.replace("${tel_1}",str(tel+1))
                    elif sheet.cell(case_id, 3).value.find("$loan_member_id") != -1:
                        row_data['data'] = sheet.cell(case_id, 3).value.replace("$loan_member_id",str(getattr(Get_data,"loan_member_id")))
                    elif sheet.cell(case_id, 3).value.find("$normal_tel") != -1:
                        row_data['data'] = sheet.cell(case_id, 3).value.replace("$normal_tel", str(getattr(Get_data,"normal_tel")))
                    elif sheet.cell(case_id, 3).value.find("$memberID") != -1:
                        row_data['data'] = sheet.cell(case_id, 3).value.replace("$memberID", str(getattr(Get_data,"memberID")))
                    else:
                        row_data['data'] = sheet.cell(case_id+1, 3).value
                    row_data['title'] = sheet.cell(case_id+1, 4).value
                    row_data['http_method'] = sheet.cell(case_id+1, 5).value
                    row_data['expected'] = sheet.cell(case_id+1, 6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
        return test_data
    def write_back(self,file_name,sheet_name,i,value,result):#专门写数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,7).value=value
        sheet.cell(i,8).value=result
        wb.save(file_name)  #保存结果
    @classmethod
    def update_tel(self,tel,filename,sheet_name):
        wb=load_workbook(filename)
        sheet=wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(filename)
if __name__ == '__main__':
    res=Do_excle().get_data(r'E:\Onlyone_pyhton\test_data\test_data.xlsx')
    #Do_excle().write_back(r'E:\Onlyone_pyhton\test_data\test_data.xlsx','loging',2,"我在写数据")
    print(res)